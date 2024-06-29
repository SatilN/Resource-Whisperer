import psutil
import time
import platform
import subprocess
from openpyxl import Workbook, load_workbook
from datetime import datetime
from ctypes import windll
import os
from flask import Flask, render_template, jsonify, send_from_directory
from threading import Thread

# Threshold values
CPU_THRESHOLD = 80.0
MEMORY_THRESHOLD = 80.0
DISK_THRESHOLD = 90.0
TEMPERATURE_THRESHOLD = 75.0  # Celsius
NETWORK_THRESHOLD = 100000000  # Example: 100 MB

# Specify the directory where the Excel file will be saved
# Excel file setup
file_name = "D:\\system_performance_log.xlsx"
directory = os.path.dirname(file_name)

# Initialize Excel workbook
try:
    workbook = load_workbook(file_name)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Timestamp", "CPU Load (%)", "Memory Usage (%)", "Disk Usage (%)", "Temperature (°C)", "Bytes Sent", "Bytes Received"])

# Flask app setup
app = Flask(__name__)

# Global variables to store the latest metrics
latest_metrics = {
    "cpu_load": 0,
    "memory_usage": 0,
    "disk_usage": 0,
    "temperature": None,
    "bytes_sent": 0,
    "bytes_recv": 0,
    "alerts": [],
    "file_location": file_name
}

def get_cpu_load():
    return psutil.cpu_percent(interval=1)

def get_memory_usage():
    mem = psutil.virtual_memory()
    return mem.percent

def get_disk_usage():
    disk = psutil.disk_usage('/')
    return disk.percent

def get_network_usage():
    net_io = psutil.net_io_counters()
    return net_io.bytes_sent, net_io.bytes_recv

def get_temperature():
    try:
        if platform.system() == "Windows":
            return get_temperature_windows()
        elif platform.system() == "Linux":
            return get_temperature_linux()
        elif platform.system() == "Darwin":
            return get_temperature_mac()
        else:
            raise NotImplementedError("Unsupported platform")
    except Exception as e:
        print(f"Error getting temperature: {e}")
        return None

def get_temperature_windows():
    try:
        import wmi

        # Check if running as administrator
        if not windll.shell32.IsUserAnAdmin():
            raise PermissionError("Administrator privileges required")

        w = wmi.WMI(namespace="root\\wmi")
        temperature_info = w.MSAcpi_ThermalZoneTemperature()
        if temperature_info:
            return temperature_info[0].CurrentTemperature / 10.0 - 273.15
        else:
            print("No temperature data available.")
            return None
    except ImportError:
        print("WMI module not installed.")
        return None
    except PermissionError as e:
        print(e)
        return None
    except Exception as e:
        print(f"Unexpected error: {e}")
        return None

def get_temperature_linux():
    try:
        temp = psutil.sensors_temperatures()
        if temp:
            for name, entries in temp.items():
                for entry in entries:
                    if entry.current:
                        return entry.current
        print("No temperature data available.")
        return None
    except Exception as e:
        print(f"Error getting temperature: {e}")
        return None

def get_temperature_mac():
    try:
        result = subprocess.run(['osx-cpu-temp'], stdout=subprocess.PIPE)
        temp_str = result.stdout.decode('utf-8')
        return float(temp_str.strip().replace('°C', ''))
    except FileNotFoundError:
        print("osx-cpu-temp not found. Please install it from https://github.com/lavoiesl/osx-cpu-temp.")
        return None
    except Exception as e:
        print(f"Error getting temperature: {e}")
        return None

def log_to_excel(cpu_load, memory_usage, disk_usage, temperature, bytes_sent, bytes_recv):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([timestamp, cpu_load, memory_usage, disk_usage, temperature, bytes_sent, bytes_recv])
    workbook.save(file_name)

def check_alerts(cpu_load, memory_usage, disk_usage, temperature, bytes_sent, bytes_recv):
    alerts = []
    if cpu_load > CPU_THRESHOLD:
        alerts.append(f"CPU Load is above {CPU_THRESHOLD}%: {cpu_load}%")
    if memory_usage > MEMORY_THRESHOLD:
        alerts.append(f"Memory Usage is above {MEMORY_THRESHOLD}%: {memory_usage}%")
    if disk_usage > DISK_THRESHOLD:
        alerts.append(f"Disk Usage is above {DISK_THRESHOLD}%: {disk_usage}%")
    if temperature and temperature > TEMPERATURE_THRESHOLD:
        alerts.append(f"Temperature is above {TEMPERATURE_THRESHOLD}°C: {temperature:.2f}°C")
    if bytes_sent > NETWORK_THRESHOLD:
        alerts.append(f"Bytes Sent is above {NETWORK_THRESHOLD}: {bytes_sent}")
    if bytes_recv > NETWORK_THRESHOLD:
        alerts.append(f"Bytes Received is above {NETWORK_THRESHOLD}: {bytes_recv}")
    return alerts

def monitor_performance(interval=5):
    while True:
        cpu_load = get_cpu_load()
        memory_usage = get_memory_usage()
        disk_usage = get_disk_usage()
        bytes_sent, bytes_recv = get_network_usage()
        temperature = get_temperature()

        log_to_excel(cpu_load, memory_usage, disk_usage, temperature, bytes_sent, bytes_recv)
        alerts = check_alerts(cpu_load, memory_usage, disk_usage, temperature, bytes_sent, bytes_recv)

        global latest_metrics
        latest_metrics = {
            "cpu_load": cpu_load,
            "memory_usage": memory_usage,
            "disk_usage": disk_usage,
            "temperature": temperature,
            "bytes_sent": bytes_sent,
            "bytes_recv": bytes_recv,
            "alerts": alerts,
            "file_location": file_name
        }

        print(f"CPU Load: {cpu_load}%")
        print(f"Memory Usage: {memory_usage}%")
        print(f"Disk Usage: {disk_usage}%")
        if temperature is not None:
            print(f"Temperature: {temperature:.2f}°C")
        else:
            print("Temperature: Not available")
        print(f"Bytes Sent: {bytes_sent}")
        print(f"Bytes Received: {bytes_recv}")
        for alert in alerts:
            print(f"ALERT: {alert}")

        print("-" * 30)
        time.sleep(interval)

@app.route('/')
def index():
    machine_name = platform.node()
    return render_template('index.html', machine_name=machine_name)

@app.route('/metrics')
def metrics():
    return jsonify(latest_metrics)

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(directory, filename)

def run_flask():
    app.run(debug=True, use_reloader=False)

if __name__ == "__main__":
    # Start the Flask app in a separate thread
    thread = Thread(target=run_flask)
    thread.start()

    # Start monitoring performance
    monitor_performance(interval=5)
